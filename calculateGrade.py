import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from getGrade import getGrade
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
load_dotenv()


submission_url = os.getenv('SUBMISSION_URL')
username = os.getenv('USERNAME')
password = os.getenv('PASSWORD')

def getStudentList():
    return list(map(lambda s: s.strip(), os.getenv('STUDENT_LIST').split(',')))

def toExcel(result):
    scope = os.getenv('SCOPE')
    file_path = 'files'
    sid_list = getStudentList()

    # 轉換 result 成 DataFrame 格式
    df = pd.DataFrame(result)

    # 1. 移除 result 不是 'correct' 的資料
    df = df[df['result'] == 'correct']

    # 2. 使用給定的 sid 順序來篩選並排序資料
    df['sid'] = pd.Categorical(df['sid'], categories=sid_list, ordered=True)
    df = df.sort_values('sid')

    # 3. 使用 groupby 來確保 sid 和 pid 的組合是唯一的，並合併 result 和 lang
    df_unique = df.groupby(['sid', 'pid'], observed=False).first().reset_index()

    # 4. 使用 pivot_table 來重塑 DataFrame 結構，使 sid 為列，pid 為欄，並包含 result 和 lang 兩個子欄位
    pivot_df = df_unique.pivot(index='sid', columns='pid', values=['result', 'lang'])

    # 使用 swaplevel 來調整欄位結構，使得 pid 作為主要欄位，result 和 lang 作為子欄位
    pivot_df = pivot_df.swaplevel(axis=1).sort_index(axis=1, level=0)

    # 填充空白值為空字串
    pivot_df.fillna('', inplace=True)

    # 5. 將結果輸出為 Excel 檔案
    output_file = f"{file_path}/{datetime.now().strftime('%m%d-%H%M%S')}-{scope}.xlsx"
    Path(file_path).mkdir(parents=True, exist_ok=True)
    pivot_df.to_excel(f'{output_file}', sheet_name='Result')

    # 6. 使用 openpyxl 來開啟生成的 Excel，並設置 "correct" 儲存格背景色為綠色
    wb = load_workbook(output_file)
    ws = wb['Result']

    # 定義綠色背景填充樣式
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # 定義字體樣式（Calibri 字體，大小 14）
    font_style = Font(name='Calibri', size=14)

    # 尋找儲存格並根據值進行格式設定
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.font = font_style  # 設定所有儲存格的字體格式
            if cell.value == 'correct':
                cell.fill = green_fill  # 設置背景色為綠色

    # 儲存 Excel 檔案
    wb.save(output_file)
    print(f"已成功輸出成 Excel 格式並設置顏色，檔案名稱為 {output_file.split('/')[-1]}")


if __name__ == '__main__':
    grades = getGrade(submission_url, username, password)
    if grades:
        # score = getScore(grades, 2)
        toExcel(grades)
        # print(score)
    else:
        print('取得成績失敗')
